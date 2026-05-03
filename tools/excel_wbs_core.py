"""Operations on `2. WBS` sheet.

Three independent operations:
  inject_hierarchy_rows  — insert L1/L2 phase/module rows above leaf tasks
  clear_junk_rows        — delete blank rows past the last data row
  add_total_row          — append a TOTAL row with SUM formulas

These run AFTER the legacy renderer when its WBS state was incomplete
(missing L1/L2 nodes) — Effort sheet VLOOKUPs depend on these rows.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .excel_workbook_core import (
    SHEET_WBS,
    WBS_COL_AI,
    WBS_COL_BA,
    WBS_COL_BE,
    WBS_COL_FE,
    WBS_COL_FEATURE,
    WBS_COL_NUM,
    WBS_COL_PM,
    WBS_COL_QC,
    WBS_COL_REFCODE,
    WBS_COL_REMARK,
    WBS_COL_TOTAL,
    WBS_DATA_START,
    WBS_HEADER_ROW,
    cell_has_value,
    clear_row,
    col_letter,
    find_last_data_row,
)


# Visual styles to match render_wbs.py conventions
_FILL_L1 = PatternFill("solid", fgColor="305496")  # dark blue
_FONT_L1 = Font(bold=True, color="FFFFFF")
_FILL_L2 = PatternFill("solid", fgColor="BDD7EE")  # light blue
_FONT_L2 = Font(bold=True)


@dataclass
class HierarchyRow:
    """A phase (L1) or module (L2) row to inject into WBS sheet."""
    code: str          # e.g. "I" or "I.A"
    feature: str       # human label
    leaf_codes: list[int] = field(default_factory=list)  # row_num values that belong under this node


@dataclass
class HierarchySpec:
    """Ordered hierarchy spec — phases each containing modules each pointing to leaf row_nums."""
    phases: list[HierarchyRow]                        # L1 nodes
    modules_by_phase: dict[str, list[HierarchyRow]]   # phase_code → L2 nodes


def _set_cell_style(ws, row: int, col: int, fill: PatternFill, font: Font) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.fill = copy(fill)
    cell.font = copy(font)


def _write_hierarchy_row(
    ws,
    row: int,
    code: str,
    feature: str,
    leaf_rows: list[int],
    is_l1: bool,
) -> None:
    """Write a phase or module row at `row` with rollup SUM formulas over its
    leaf rows in cols F (Total), G (BE), H (FE), I (BA), J (QC), K (PM)."""
    clear_row(ws, row, max_col=WBS_COL_REMARK)

    fill = _FILL_L1 if is_l1 else _FILL_L2
    font = _FONT_L1 if is_l1 else _FONT_L2
    for c in range(1, WBS_COL_REMARK + 1):
        _set_cell_style(ws, row, c, fill, font)

    ws.cell(row, WBS_COL_NUM).value = code
    ws.cell(row, WBS_COL_FEATURE).value = feature

    if leaf_rows:
        for col in (WBS_COL_TOTAL, WBS_COL_BE, WBS_COL_FE, WBS_COL_AI,
                    WBS_COL_BA, WBS_COL_QC, WBS_COL_PM):
            cl = get_column_letter(col)
            refs = ",".join(f"{cl}{r}" for r in leaf_rows)
            ws.cell(row, col).value = f"=SUM({refs})"


def _gather_leaf_index(ws, last_row: int) -> dict[int, int]:
    """Return mapping `row_num value (col B) → spreadsheet row`, only for rows
    where col B is numeric (leaf tasks)."""
    out: dict[int, int] = {}
    for r in range(WBS_DATA_START, last_row + 1):
        v = ws.cell(r, WBS_COL_NUM).value
        if isinstance(v, (int, float)):
            out[int(v)] = r
    return out


def inject_hierarchy_rows(
    wb: Workbook,
    spec: HierarchySpec,
) -> dict:
    """Insert L1/L2 rows ABOVE their leaf tasks. Idempotent — skips codes that
    already exist as a row's col-B value.

    Strategy: insert at the row position right BEFORE the first leaf belonging
    to the phase/module. Subsequent leaf row numbers shift down and we rebuild
    the leaf index between insertions.
    """
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    # Existing string codes (already-injected hierarchy)
    existing_codes = set()
    last = find_last_data_row(
        ws, start_row=WBS_DATA_START,
        check_cols=(WBS_COL_NUM, WBS_COL_FEATURE),
    )
    for r in range(WBS_DATA_START, last + 1):
        v = ws.cell(r, WBS_COL_NUM).value
        if isinstance(v, str):
            existing_codes.add(v.strip())

    inserted: list[dict] = []

    for phase in spec.phases:
        # Find phase's leaf rows (union of its module's leaf_codes)
        all_leafs = list(phase.leaf_codes)
        for mod in spec.modules_by_phase.get(phase.code, []):
            all_leafs.extend(mod.leaf_codes)

        if not all_leafs:
            continue

        # 1. Inject phase row above first leaf
        if phase.code not in existing_codes:
            leaf_idx = _gather_leaf_index(ws, ws.max_row)
            leaf_rows = sorted(leaf_idx[c] for c in all_leafs if c in leaf_idx)
            if not leaf_rows:
                continue
            insert_at = leaf_rows[0]
            ws.insert_rows(insert_at, amount=1)

            # After insert, leaf rows have shifted by +1; rebuild for formula refs
            leaf_idx = _gather_leaf_index(ws, ws.max_row)
            leaf_rows_refreshed = sorted(leaf_idx[c] for c in all_leafs if c in leaf_idx)
            _write_hierarchy_row(
                ws, insert_at,
                code=phase.code, feature=phase.feature,
                leaf_rows=leaf_rows_refreshed, is_l1=True,
            )
            existing_codes.add(phase.code)
            inserted.append({"code": phase.code, "row": insert_at, "level": 1})

        # 2. Inject each module row above its first leaf
        for mod in spec.modules_by_phase.get(phase.code, []):
            if mod.code in existing_codes:
                continue
            if not mod.leaf_codes:
                continue
            leaf_idx = _gather_leaf_index(ws, ws.max_row)
            mod_rows = sorted(leaf_idx[c] for c in mod.leaf_codes if c in leaf_idx)
            if not mod_rows:
                continue
            insert_at = mod_rows[0]
            ws.insert_rows(insert_at, amount=1)

            leaf_idx = _gather_leaf_index(ws, ws.max_row)
            mod_rows_refreshed = sorted(leaf_idx[c] for c in mod.leaf_codes if c in leaf_idx)
            _write_hierarchy_row(
                ws, insert_at,
                code=mod.code, feature=mod.feature,
                leaf_rows=mod_rows_refreshed, is_l1=False,
            )
            existing_codes.add(mod.code)
            inserted.append({"code": mod.code, "row": insert_at, "level": 2})

    return {"inserted": inserted, "count": len(inserted)}


def clear_junk_rows(wb: Workbook) -> dict:
    """Delete every row past the last meaningful row in `2. WBS`."""
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    last = find_last_data_row(
        ws, start_row=WBS_DATA_START,
        check_cols=(WBS_COL_NUM, WBS_COL_REFCODE, WBS_COL_FEATURE),
    )
    target_max = last + 1  # keep one buffer row for the future TOTAL

    if ws.max_row <= target_max:
        return {"removed": 0, "kept_until_row": last}

    delete_count = ws.max_row - target_max
    ws.delete_rows(target_max + 1, delete_count)

    return {"removed": delete_count, "kept_until_row": last}


def add_total_row(wb: Workbook, label: str = "TOTAL") -> dict:
    """Append a TOTAL row with SUM formulas that sum only L1 phase rows.

    Why only L1 rows: leaves and L2 modules are already aggregated into L1
    rollups; summing all rows would double-count.
    """
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    last = find_last_data_row(
        ws, start_row=WBS_DATA_START,
        check_cols=(WBS_COL_NUM, WBS_COL_REFCODE, WBS_COL_FEATURE),
    )
    if last < WBS_DATA_START:
        return {"row": None, "skipped": "no data"}

    # Already has a TOTAL row?
    for r in range(WBS_DATA_START, last + 1):
        v = ws.cell(r, WBS_COL_FEATURE).value
        if isinstance(v, str) and v.strip().upper() == label.upper():
            return {"row": r, "skipped": "already exists"}

    # Find L1 rows (col B is a string without a dot)
    l1_rows: list[int] = []
    for r in range(WBS_DATA_START, last + 1):
        v = ws.cell(r, WBS_COL_NUM).value
        if isinstance(v, str) and v.strip() and "." not in v:
            l1_rows.append(r)

    target_row = last + 1
    clear_row(ws, target_row, max_col=WBS_COL_REMARK)
    ws.cell(target_row, WBS_COL_FEATURE).value = label

    fill = _FILL_L1
    font = _FONT_L1
    for c in range(1, WBS_COL_REMARK + 1):
        _set_cell_style(ws, target_row, c, fill, font)

    if l1_rows:
        for col in (WBS_COL_TOTAL, WBS_COL_BE, WBS_COL_FE, WBS_COL_AI,
                    WBS_COL_BA, WBS_COL_QC, WBS_COL_PM):
            cl = get_column_letter(col)
            refs = ",".join(f"{cl}{r}" for r in l1_rows)
            ws.cell(target_row, col).value = f"=SUM({refs})"

    return {"row": target_row, "summed_rows": l1_rows}


def consolidate_hierarchy(wb: Workbook) -> dict:
    """Remove duplicate numeric-col-B leaf rows when L1/L2/L3 hierarchy with
    effort already exists. After removal, the hierarchy block shifts to start
    at WBS_DATA_START (row 9). Idempotent.

    Used when the renderer wrote both flat L4 tasks AND a hierarchical breakdown
    that effectively duplicates the same work — keep the hierarchy, drop the
    flat leaves.
    """
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    leaf_rows: list[int] = []
    hier_rows: list[int] = []
    last = find_last_data_row(
        ws, start_row=WBS_DATA_START,
        check_cols=(WBS_COL_NUM, WBS_COL_REFCODE, WBS_COL_FEATURE),
    )
    for r in range(WBS_DATA_START, last + 1):
        b = ws.cell(r, WBS_COL_NUM).value
        if isinstance(b, (int, float)):
            leaf_rows.append(r)
        elif isinstance(b, str) and b.strip():
            hier_rows.append(r)

    # Only drop leaves if hierarchy block exists AND has effort numbers
    has_effort = any(
        isinstance(ws.cell(r, WBS_COL_BE).value, (int, float))
        for r in hier_rows
    )
    if not hier_rows or not has_effort:
        return {"action": "skipped", "reason": "no hierarchy with effort to keep"}

    for r in sorted(leaf_rows, reverse=True):
        ws.delete_rows(r)

    return {
        "removed_leaf_rows": len(leaf_rows),
        "kept_hierarchy_rows": len(hier_rows),
    }


def rebuild_l1_l2_rollups(wb: Workbook) -> dict:
    """For each L1 / L2 row in WBS, rewrite cols F-K as SUM formulas over the
    immediate children rows (L2 children of L1, L3 children of L2).

    Assumes WBS has been consolidated (no duplicate leaf rows). Reads col B
    codes to figure out parent → child relationships.
    """
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    last = find_last_data_row(
        ws, start_row=WBS_DATA_START,
        check_cols=(WBS_COL_NUM, WBS_COL_FEATURE),
    )

    # Map code → row
    by_code: dict[str, int] = {}
    for r in range(WBS_DATA_START, last + 1):
        v = ws.cell(r, WBS_COL_NUM).value
        if isinstance(v, str) and v.strip():
            by_code[v.strip()] = r

    rebuilt = 0
    for code, r in by_code.items():
        depth = code.count(".") + 1

        if depth >= 3:
            # L3 is the leaf in consolidated view.
            # Total = SUM(BE:PM for this row); BA/QC/PM use master-data pct × (BE+FE+AI).
            be_col = col_letter(WBS_COL_BE)   # G
            fe_col = col_letter(WBS_COL_FE)   # H
            ai_col = col_letter(WBS_COL_AI)   # I
            pm_col = col_letter(WBS_COL_PM)   # L
            be = ws.cell(r, WBS_COL_BE).value
            fe = ws.cell(r, WBS_COL_FE).value
            has_effort = isinstance(be, (int, float)) or isinstance(fe, (int, float))
            # Total = SUM(G:L) = BE+FE+AI+BA+QC+PM
            ws.cell(r, WBS_COL_TOTAL).value = f"=SUM({be_col}{r}:{pm_col}{r})"
            if has_effort:
                dev = f"({be_col}{r}+{fe_col}{r}+{ai_col}{r})"
                ws.cell(r, WBS_COL_BA).value = f"={dev}*pct_ba"
                ws.cell(r, WBS_COL_QC).value = f"={dev}*pct_qc"
                ws.cell(r, WBS_COL_PM).value = f"={dev}*pct_pm"
            else:
                for col in (WBS_COL_BA, WBS_COL_QC, WBS_COL_PM):
                    cell = ws.cell(r, col)
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = None
            rebuilt += 1
            continue

        # L1/L2: SUM over immediate children (one level deeper)
        prefix = code + "."
        target_depth = depth + 1
        child_rows: list[int] = []
        for child_code, child_row in by_code.items():
            if not child_code.startswith(prefix):
                continue
            if child_code.count(".") + 1 != target_depth:
                continue
            child_rows.append(child_row)

        if not child_rows:
            continue
        child_rows.sort()

        for col in (WBS_COL_TOTAL, WBS_COL_BE, WBS_COL_FE, WBS_COL_AI,
                    WBS_COL_BA, WBS_COL_QC, WBS_COL_PM):
            cl = get_column_letter(col)
            refs = ",".join(f"{cl}{cr}" for cr in child_rows)
            ws.cell(r, col).value = f"=SUM({refs})"
        rebuilt += 1

    return {"rebuilt_rows": rebuilt}


def insert_ai_column_wbs(wb: Workbook) -> dict:
    """Insert the AI/ML (MD) column into `2. WBS` between FE (H) and BA (old I).

    After this call the column layout is:
      G=BE, H=FE, I=AI/ML (new, blank), J=BA (was I), K=QC (was J),
      L=PM (was K), M=Remark (was L).

    Call `rebuild_l1_l2_rollups(wb)` immediately after to rewrite all
    formulas with the new column positions.
    """
    if SHEET_WBS not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_WBS}")
    ws = wb[SHEET_WBS]

    # Idempotent: skip if the header at col WBS_COL_AI already says AI/ML
    existing_header = ws.cell(WBS_HEADER_ROW, WBS_COL_AI).value
    if isinstance(existing_header, str) and "AI" in existing_header.upper():
        return {"action": "skipped", "reason": "AI column already present"}

    ws.insert_cols(WBS_COL_AI)  # shifts BA→J, QC→K, PM→L, Remark→M

    # Write header at the new AI column position
    ws.cell(WBS_HEADER_ROW, WBS_COL_AI).value = "AI/ML (MD)"

    return {"action": "inserted", "col": WBS_COL_AI, "header": "AI/ML (MD)"}


def derive_hierarchy_from_state(state_phases: list[dict]) -> HierarchySpec:
    """Convert a list of phase dicts into a HierarchySpec.

    Schema for `state_phases`:
      [
        {"code": "I", "feature": "Setup",
         "modules": [
           {"code": "I.A", "feature": "Init",
            "leaf_codes": [1, 2, 3]}
         ]},
        ...
      ]
    """
    phases: list[HierarchyRow] = []
    modules_by_phase: dict[str, list[HierarchyRow]] = {}

    for ph in state_phases:
        phase = HierarchyRow(
            code=ph["code"], feature=ph.get("feature", ""),
            leaf_codes=ph.get("leaf_codes", []),
        )
        mods: list[HierarchyRow] = []
        for m in ph.get("modules", []):
            mods.append(HierarchyRow(
                code=m["code"],
                feature=m.get("feature", ""),
                leaf_codes=m.get("leaf_codes", []),
            ))
        phases.append(phase)
        modules_by_phase[phase.code] = mods

    return HierarchySpec(phases=phases, modules_by_phase=modules_by_phase)
