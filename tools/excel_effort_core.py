"""Operations on `1. Effort` sheet.

Column layout (after AI column was inserted into WBS):
  B=module_id, C=name(VLOOKUP), D=Total_MD,
  E=BE_MD, F=FE_MD, G=AI_MD, H=RA_MD, I=Test_MD, J=PM_MD,
  K=Total_USD,
  L=BE_USD, M=FE_USD, N=AI_USD, O=RA_USD, P=Test_USD, Q=PM_USD

PM_MD and Test_MD (QC) are formula-derived from master-data pct:
  PM_MD  = pct_pm  × (BE + FE + AI)
  Test_MD = pct_qc × (BE + FE + AI)
Only BA (RA_MD) and the dev columns are pulled from WBS via VLOOKUP.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .excel_workbook_core import (
    EFFORT_DATA_START,
    SHEET_EFFORT,
    cell_has_value,
    clear_row,
    find_last_data_row,
)


# ── Effort sheet column map (1-indexed) ──────────────────────────────────────
E_COL_HASH = 2      # B — module ID (I, I.A, ...)
E_COL_NAME = 3      # C — module name (VLOOKUP from WBS)
E_COL_TOTAL_MD = 4  # D
E_COL_BE_MD = 5     # E
E_COL_FE_MD = 6     # F
E_COL_AI_MD = 7     # G  (AI/ML effort — new)
E_COL_RA_MD = 8     # H  (Requirements Analysis / BA)
E_COL_TEST_MD = 9   # I  (Testing / QC  — formula: pct_qc × dev)
E_COL_PM_MD = 10    # J  (Project Mgmt  — formula: pct_pm × dev)
E_COL_TOTAL_USD = 11  # K
E_COL_BE_USD = 12   # L
E_COL_FE_USD = 13   # M
E_COL_AI_USD = 14   # N
E_COL_RA_USD = 15   # O
E_COL_TEST_USD = 16  # P
E_COL_PM_USD = 17   # Q

# WBS lookup range (cols B..M = 13 cols after AI col inserted)
WBS_LOOKUP_RANGE = "'2. WBS'!$B$5:$M$9999"

# VLOOKUP column indices (relative to lookup range starting at col B):
#   index 3  = D = feature name
#   index 6  = G = BE
#   index 7  = H = FE
#   index 8  = I = AI (new)
#   index 9  = J = BA
_VLOOKUP_FEATURE = 3
_VLOOKUP_BE = 6
_VLOOKUP_FE = 7
_VLOOKUP_AI = 8
_VLOOKUP_BA = 9

# Standard header labels for row 5
EFFORT_HEADERS = {
    E_COL_HASH:      "#",
    E_COL_NAME:      "Module / Feature",
    E_COL_TOTAL_MD:  "Total (MD)",
    E_COL_BE_MD:     "BE Coding (MD)",
    E_COL_FE_MD:     "FE/Mobile Coding (MD)",
    E_COL_AI_MD:     "AI/ML (MD)",
    E_COL_RA_MD:     "Requirement Analysis (MD)",
    E_COL_TEST_MD:   "Testing (MD)",
    E_COL_PM_MD:     "Project Management (MD)",
    E_COL_TOTAL_USD: "Total (USD)",
    E_COL_BE_USD:    "BE Coding (USD)",
    E_COL_FE_USD:    "FE/Mobile (USD)",
    E_COL_AI_USD:    "AI/ML (USD)",
    E_COL_RA_USD:    "Requirement Analysis (USD)",
    E_COL_TEST_USD:  "Testing (USD)",
    E_COL_PM_USD:    "Project Management (USD)",
}

_EFFORT_HEADER_ROW = 5
_MAX_COL = E_COL_PM_USD  # rightmost column we manage (17)


def write_effort_headers(wb: Workbook) -> None:
    """Write standard column headers at row 5 of the Effort sheet."""
    if SHEET_EFFORT not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_EFFORT}")
    ws = wb[SHEET_EFFORT]
    for col, label in EFFORT_HEADERS.items():
        ws.cell(_EFFORT_HEADER_ROW, col).value = label


def rebuild_total_row(wb: Workbook) -> dict:
    """Replace the TOTAL row so it sums ALL L1 module rows (not just 3)."""
    if SHEET_EFFORT not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_EFFORT}")
    ws = wb[SHEET_EFFORT]

    total_row = None
    for r in range(EFFORT_DATA_START, ws.max_row + 1):
        v = ws.cell(r, E_COL_NAME).value
        if isinstance(v, str) and v.strip().upper() == "TOTAL":
            total_row = r
            break

    if total_row is None:
        last = find_last_data_row(ws, start_row=EFFORT_DATA_START,
                                  check_cols=(E_COL_HASH, E_COL_NAME))
        if last < EFFORT_DATA_START:
            return {"row": None, "skipped": "no data"}
        total_row = last + 1
        clear_row(ws, total_row, max_col=_MAX_COL)
        ws.cell(total_row, E_COL_NAME).value = "TOTAL"

    # Sum only L1 rows (no dot in code)
    l1_rows: list[int] = []
    for r in range(EFFORT_DATA_START, total_row):
        v = ws.cell(r, E_COL_HASH).value
        if isinstance(v, str) and v.strip() and "." not in v:
            l1_rows.append(r)

    if not l1_rows:
        return {"row": total_row, "skipped": "no L1 rows above"}

    for col in range(E_COL_TOTAL_MD, _MAX_COL + 1):
        cl = get_column_letter(col)
        refs = ",".join(f"{cl}{r}" for r in l1_rows)
        ws.cell(total_row, col).value = f"=SUM({refs})"

    return {"row": total_row, "summed_rows": l1_rows}


def rebuild_for_modules(wb: Workbook, module_codes: list[str]) -> dict:
    """Rewrite Effort rows so col B = canonical WBS module codes.

    For each module row:
    - Col C: VLOOKUP feature name from WBS.
    - Cols E,F,G (BE/FE/AI MD): VLOOKUP from WBS.
    - Col H (RA/BA MD): VLOOKUP from WBS (explicit analysis effort).
    - Col I (Test/QC MD): formula = pct_qc × (BE+FE+AI).
    - Col J (PM MD):     formula = pct_pm × (BE+FE+AI).
    - USD cols: each MD col × named rate (rate_dev / rate_ai / rate_ba / rate_qc / rate_pm).
    """
    if SHEET_EFFORT not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_EFFORT}")
    ws = wb[SHEET_EFFORT]

    # Locate or clear existing TOTAL row
    existing_total_row = None
    for r in range(EFFORT_DATA_START, ws.max_row + 1):
        v = ws.cell(r, E_COL_NAME).value
        if isinstance(v, str) and v.strip().upper() == "TOTAL":
            existing_total_row = r
            break

    n = len(module_codes)
    new_total_row = EFFORT_DATA_START + n

    if existing_total_row is not None and existing_total_row != new_total_row:
        clear_row(ws, existing_total_row, max_col=_MAX_COL)

    for i, code in enumerate(module_codes):
        r = EFFORT_DATA_START + i
        clear_row(ws, r, max_col=_MAX_COL)

        # Module ID
        ws.cell(r, E_COL_HASH).value = code
        # Feature name via VLOOKUP
        ws.cell(r, E_COL_NAME).value = (
            f"=VLOOKUP($B{r},{WBS_LOOKUP_RANGE},{_VLOOKUP_FEATURE},FALSE)"
        )

        # Dev effort from WBS
        ws.cell(r, E_COL_BE_MD).value = (
            f"=VLOOKUP($B{r},{WBS_LOOKUP_RANGE},{_VLOOKUP_BE},FALSE)"
        )
        ws.cell(r, E_COL_FE_MD).value = (
            f"=VLOOKUP($B{r},{WBS_LOOKUP_RANGE},{_VLOOKUP_FE},FALSE)"
        )
        ws.cell(r, E_COL_AI_MD).value = (
            f"=VLOOKUP($B{r},{WBS_LOOKUP_RANGE},{_VLOOKUP_AI},FALSE)"
        )
        # Requirements Analysis (BA) from WBS
        ws.cell(r, E_COL_RA_MD).value = (
            f"=VLOOKUP($B{r},{WBS_LOOKUP_RANGE},{_VLOOKUP_BA},FALSE)"
        )

        # Testing (QC) and PM derived from master-data pct × dev
        dev = f"(E{r}+F{r}+G{r})"  # BE + FE + AI
        ws.cell(r, E_COL_TEST_MD).value = f"={dev}*pct_qc"
        ws.cell(r, E_COL_PM_MD).value = f"={dev}*pct_pm"

        # Total MD
        ws.cell(r, E_COL_TOTAL_MD).value = f"=SUM(E{r}:J{r})"

        # USD columns — rates are USD/month; divide by 20 working days/month
        ws.cell(r, E_COL_BE_USD).value = f"=E{r}*(rate_dev/20)"
        ws.cell(r, E_COL_FE_USD).value = f"=F{r}*(rate_dev/20)"
        ws.cell(r, E_COL_AI_USD).value = f"=G{r}*(rate_ai/20)"
        ws.cell(r, E_COL_RA_USD).value = f"=H{r}*(rate_ba/20)"
        ws.cell(r, E_COL_TEST_USD).value = f"=I{r}*(rate_qc/20)"
        ws.cell(r, E_COL_PM_USD).value = f"=J{r}*(rate_pm/20)"
        ws.cell(r, E_COL_TOTAL_USD).value = f"=SUM(L{r}:Q{r})"

    # Clear leftover rows
    for r in range(new_total_row + 1, max(new_total_row + 10, ws.max_row + 1)):
        if not cell_has_value(ws, r, E_COL_HASH) and not cell_has_value(ws, r, E_COL_NAME):
            break
        clear_row(ws, r, max_col=_MAX_COL)

    # Write fresh TOTAL row
    clear_row(ws, new_total_row, max_col=_MAX_COL)
    ws.cell(new_total_row, E_COL_NAME).value = "TOTAL"
    l1_rows = [EFFORT_DATA_START + i for i, c in enumerate(module_codes) if "." not in c]
    if l1_rows:
        for col in range(E_COL_TOTAL_MD, _MAX_COL + 1):
            cl = get_column_letter(col)
            refs = ",".join(f"{cl}{r}" for r in l1_rows)
            ws.cell(new_total_row, col).value = f"=SUM({refs})"

    return {"module_count": n, "total_row": new_total_row, "l1_rows": l1_rows}
