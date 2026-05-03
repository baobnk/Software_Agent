"""Operations on `3. Delivery Plan` sheet.

Three writes:
  write_master_planning(...)         — fill Module rows 7..17 (Start, End, sprint marks)
  write_resource_planning(...)       — fill rows 20..N for each role; rebuild TOTAL
  write_deliverable_milestones(...)  — fill rows 30..34 with confirmed dates
"""
from __future__ import annotations

from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .excel_workbook_core import (
    SHEET_DELIVERY,
    cell_has_value,
    clear_row,
)


def _safe_set(ws, row: int, col: int, value=None, fill=None) -> None:
    """Skip merged-slave cells silently."""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    if value is not None or value == 0:
        cell.value = value
    if fill is not None:
        cell.fill = fill


# Delivery sheet layout
DP_HEADER_ROW = 5
DP_DATA_START = 7        # First module row (row 6 = REQUIREMENT GATHERING static)
DP_DATA_MAX = 21         # Last module row (15 slots: I, I.A, I.B, II, II.A–II.G, III, III.A–III.C)

DP_RP_HEADER_ROW = 23
DP_RP_DATA_START = 24    # First role row (7 roles: PM, TL, Dev, BA, QC, Designer, DevOps)

DP_DM_HEADER_ROW = 33
DP_DM_DATA_START = 34    # First deliverable row (5 milestones)
DP_DM_DATA_MAX = 38

DP_COL_NUM = 2       # B
DP_COL_NAME = 3      # C
DP_COL_START = 4     # D
DP_COL_END = 5       # E
DP_COL_FIRST_WEEK = 6  # F = W1, G = W2, ...

# Visual fill for an active sprint cell
_FILL_ACTIVE = PatternFill("solid", fgColor="92D050")  # green
_FILL_UAT = PatternFill("solid", fgColor="9DC3E6")     # light blue
_FILL_NURSING = PatternFill("solid", fgColor="FFD966") # amber


def _to_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v).date()
        except Exception:
            return None
    return None


def write_master_planning(wb: Workbook, gantt: dict) -> dict:
    """Fill master planning rows from gantt['modules'].

    Each module row matches the template by col B (module ID like "I", "I.A").
    Writes Start (col D), End (col E), and active sprint cells (col F+).
    """
    if SHEET_DELIVERY not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_DELIVERY}")
    ws = wb[SHEET_DELIVERY]

    # Index template rows by their existing module ID (col B)
    by_code: dict[str, int] = {}
    for r in range(DP_DATA_START, DP_DATA_MAX + 1):
        v = ws.cell(r, DP_COL_NUM).value
        if isinstance(v, str) and v.strip():
            by_code[v.strip()] = r

    written = 0
    for mod in gantt["modules"]:
        r = by_code.get(mod["code"])
        if r is None:
            # Append below existing rows if no template slot
            r = max(DP_DATA_START, max(by_code.values()) + 1) if by_code else DP_DATA_START
            ws.cell(r, DP_COL_NUM).value = mod["code"]
            by_code[mod["code"]] = r

        # Clear sprint cells first (so this is idempotent)
        empty_fill = PatternFill(fill_type=None)
        for c in range(DP_COL_FIRST_WEEK, DP_COL_FIRST_WEEK + gantt["total_weeks"]):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = empty_fill

        start_cell = ws.cell(r, DP_COL_START)
        if not isinstance(start_cell, MergedCell):
            start_cell.value = _to_date(mod["start"]) or mod["start"]
        end_cell = ws.cell(r, DP_COL_END)
        if not isinstance(end_cell, MergedCell):
            end_cell.value = _to_date(mod["end"]) or mod["end"]

        # Mark active weeks
        for w in range(mod["week_start"], mod["week_end"] + 1):
            col = DP_COL_FIRST_WEEK + (w - 1)
            cell = ws.cell(r, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = 1
            cell.fill = _FILL_ACTIVE

        written += 1

    return {"rows_written": written, "total_weeks": gantt["total_weeks"]}


def _unmerge_overlapping(ws, row: int, col_min: int, col_max: int) -> None:
    """Unmerge any range that touches `row` between col_min..col_max."""
    to_unmerge = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col_max and rng.max_col >= col_min:
            to_unmerge.append(str(rng))
    for rng_str in to_unmerge:
        ws.unmerge_cells(rng_str)


def write_resource_planning(
    wb: Workbook,
    allocation: dict,
    gantt: dict,
    deadline_date: Optional[date] = None,
) -> dict:
    """Replace Resource Planning rows (20..N) with per-role × per-week alloc.

    - Each role gets one row.
    - Cols D, E are Start/End for that role's involvement (first/last
      non-zero week translated back to a date via gantt anchor_monday).
    - Cols F..F+total_weeks-1 are weekly allocation values.
    - The TOTAL row goes immediately under the last role.
    """
    if SHEET_DELIVERY not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_DELIVERY}")
    ws = wb[SHEET_DELIVERY]

    total_weeks = gantt["total_weeks"]

    # Discover existing TOTAL row to determine clear range
    rp_total_row = None
    for r in range(DP_RP_DATA_START, DP_RP_DATA_START + 30):
        v = ws.cell(r, DP_COL_NAME).value
        if isinstance(v, str) and v.strip().upper() == "TOTAL":
            rp_total_row = r
            break

    # Clear the entire RP block — also unmerge anything in our footprint
    n_roles = len(allocation["roles"])
    end_clear = max(rp_total_row or 0, DP_RP_DATA_START + n_roles + 1)
    max_write_col = DP_COL_FIRST_WEEK + total_weeks + 1
    for r in range(DP_RP_DATA_START, end_clear + 1):
        _unmerge_overlapping(ws, r, 2, max_write_col)
        clear_row(ws, r, max_col=max_write_col + 5)

    # Anchor monday for date math
    anchor_str = gantt.get("anchor_monday")
    if anchor_str:
        anchor = datetime.fromisoformat(anchor_str).date()
    else:
        anchor = None

    roles = allocation["roles"]
    seq = 0
    last_role_row = DP_RP_DATA_START - 1
    for role_name, weekly in roles.items():
        # Normalize week keys to int — JSON round-trips turn them into strings
        weekly = {int(k): v for k, v in weekly.items()}
        seq += 1
        r = DP_RP_DATA_START + seq - 1
        last_role_row = r
        for col, val in [(DP_COL_NUM, seq), (DP_COL_NAME, role_name)]:
            cell = ws.cell(r, col)
            if not isinstance(cell, MergedCell):
                cell.value = val

        active_weeks = sorted(w for w, v in weekly.items() if v and v > 0)
        if active_weeks and anchor:
            from datetime import timedelta
            start = anchor + timedelta(weeks=active_weeks[0] - 1)
            end = anchor + timedelta(weeks=active_weeks[-1] - 1, days=4)
            if deadline_date and end > deadline_date:
                end = deadline_date
            for col, val in [(DP_COL_START, start), (DP_COL_END, end)]:
                cell = ws.cell(r, col)
                if not isinstance(cell, MergedCell):
                    cell.value = val
        elif active_weeks:
            for col in (DP_COL_START, DP_COL_END):
                cell = ws.cell(r, col)
                if not isinstance(cell, MergedCell):
                    cell.value = "TBD"

        for w in range(1, total_weeks + 1):
            v = weekly.get(w, 0)
            col = DP_COL_FIRST_WEEK + (w - 1)
            if v and v > 0:
                cell = ws.cell(r, col)
                if not isinstance(cell, MergedCell):
                    cell.value = v

    # TOTAL row
    total_row = last_role_row + 1
    clear_row(ws, total_row, max_col=DP_COL_FIRST_WEEK + total_weeks + 5)
    name_cell = ws.cell(total_row, DP_COL_NAME)
    if not isinstance(name_cell, MergedCell):
        name_cell.value = "TOTAL"
    for w in range(1, total_weeks + 1):
        col = DP_COL_FIRST_WEEK + (w - 1)
        cl = get_column_letter(col)
        cell = ws.cell(total_row, col)
        if not isinstance(cell, MergedCell):
            cell.value = f"=SUM({cl}{DP_RP_DATA_START}:{cl}{last_role_row})"

    return {
        "roles_written": len(roles),
        "total_row": total_row,
        "total_weeks": total_weeks,
    }


def write_deliverable_milestones(wb: Workbook, milestones: list[dict]) -> dict:
    """Write/overwrite milestone rows 30.. with name, seq, Start, End, Deliverable.

    Always overwrites — use confirmed milestone list from HITL.
    """
    if SHEET_DELIVERY not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_DELIVERY}")
    ws = wb[SHEET_DELIVERY]

    written = 0
    for i, ms in enumerate(milestones):
        r = DP_DM_DATA_START + i
        if r > DP_DM_DATA_MAX + 5:
            break

        _unmerge_overlapping(ws, r, 2, 8)

        for col, val in [
            (DP_COL_NUM, ms.get("seq", i + 1)),
            (DP_COL_NAME, ms["name"]),
            (DP_COL_START, _to_date(ms["start"]) or ms["start"]),
            (DP_COL_END, _to_date(ms["end"]) or ms["end"]),
        ]:
            cell = ws.cell(r, col)
            if not isinstance(cell, MergedCell):
                cell.value = val

        # Deliverables col F (col 6)
        if ms.get("deliverable"):
            deliv_cell = ws.cell(r, 6)
            if not isinstance(deliv_cell, MergedCell):
                deliv_cell.value = ms["deliverable"]

        written += 1

    return {"milestones_written": written}
