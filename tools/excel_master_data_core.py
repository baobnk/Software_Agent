"""Master Data sheet operations.

Adds new roles (e.g. AI Engineer) into the two sub-tables on `4. Master Data`:
- "Percent on Dev" block (rows 4..7 by default) — appended below the last role
- "Rate" block (rows 11..14) — appended below the last role
And creates corresponding workbook-scoped named ranges (e.g. pct_ai, rate_ai).
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .excel_workbook_core import (
    SHEET_MASTER,
    cell_has_value,
    upsert_named_range,
)


PCT_BLOCK_HEADER_ROW = 3
PCT_BLOCK_FIRST_DATA_ROW = 4
PCT_BLOCK_MAX_ROW = 8        # leave row 9 (OB header) reserved
RATE_BLOCK_HEADER_ROW = 10
RATE_BLOCK_FIRST_DATA_ROW = 11
RATE_BLOCK_MAX_ROW = 30      # arbitrary safe cap


def _find_pct_insert_row(ws) -> int:
    """First empty row in the Percent block (below last role with a label)."""
    last = PCT_BLOCK_HEADER_ROW
    for r in range(PCT_BLOCK_FIRST_DATA_ROW, PCT_BLOCK_MAX_ROW + 1):
        v = ws.cell(r, 2).value  # col B = label
        if isinstance(v, str) and v.strip() and v.strip() != "OB":
            last = r
    return last + 1


def _find_rate_insert_row(ws) -> int:
    last = RATE_BLOCK_HEADER_ROW
    for r in range(RATE_BLOCK_FIRST_DATA_ROW, RATE_BLOCK_MAX_ROW + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip():
            last = r
    return last + 1


def upsert_role(
    wb: Workbook,
    role_label: str,
    pct_on_dev: float,
    rate_usd: float,
    pct_named_range: str,
    rate_named_range: str,
    remark: str = "",
) -> dict:
    """Append the role to both sub-tables; create/update named ranges.

    Returns: {pct_row, rate_row, pct_cell, rate_cell}
    """
    if SHEET_MASTER not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_MASTER}")
    ws = wb[SHEET_MASTER]

    # Skip if role already present in pct block
    pct_row = None
    for r in range(PCT_BLOCK_FIRST_DATA_ROW, PCT_BLOCK_MAX_ROW + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower() == role_label.lower():
            pct_row = r
            break
    if pct_row is None:
        pct_row = _find_pct_insert_row(ws)
        ws.cell(pct_row, 2).value = role_label
        if remark:
            ws.cell(pct_row, 4).value = remark
    ws.cell(pct_row, 3).value = pct_on_dev

    # Same for rate block
    rate_row = None
    for r in range(RATE_BLOCK_FIRST_DATA_ROW, RATE_BLOCK_MAX_ROW + 1):
        v = ws.cell(r, 2).value
        if isinstance(v, str) and v.strip().lower() == role_label.lower():
            rate_row = r
            break
    if rate_row is None:
        rate_row = _find_rate_insert_row(ws)
        ws.cell(rate_row, 2).value = role_label
    ws.cell(rate_row, 3).value = rate_usd

    pct_cell = f"$C${pct_row}"
    rate_cell = f"$C${rate_row}"
    upsert_named_range(wb, pct_named_range, SHEET_MASTER, pct_cell)
    upsert_named_range(wb, rate_named_range, SHEET_MASTER, rate_cell)

    return {
        "pct_row": pct_row,
        "rate_row": rate_row,
        "pct_cell": f"'{SHEET_MASTER}'!{pct_cell}",
        "rate_cell": f"'{SHEET_MASTER}'!{rate_cell}",
    }


def update_role_rates(wb: Workbook, rates: dict[str, float]) -> dict:
    """Update rate values for existing roles in the Rate block.

    Args:
        rates: {role_label: new_rate_usd}, e.g. {"PM": 2500, "Developer": 2500}.

    Returns: {role: {old_rate, new_rate, row}} for each updated role.
    """
    if SHEET_MASTER not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_MASTER}")
    ws = wb[SHEET_MASTER]

    updated: dict = {}
    for r in range(RATE_BLOCK_FIRST_DATA_ROW, RATE_BLOCK_MAX_ROW + 1):
        label = ws.cell(r, 2).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        if label in rates:
            old = ws.cell(r, 3).value
            ws.cell(r, 3).value = rates[label]
            updated[label] = {"old_rate": old, "new_rate": rates[label], "row": r}

    return updated


def update_role_pcts(wb: Workbook, pcts: dict[str, float]) -> dict:
    """Update percentage values for existing roles in the Percent block.

    Args:
        pcts: {role_label: new_pct}, e.g. {"PM": 0.05, "QC": 0.1}.
    """
    if SHEET_MASTER not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_MASTER}")
    ws = wb[SHEET_MASTER]

    updated: dict = {}
    for r in range(PCT_BLOCK_FIRST_DATA_ROW, PCT_BLOCK_MAX_ROW + 1):
        label = ws.cell(r, 2).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        if label in pcts:
            old = ws.cell(r, 3).value
            ws.cell(r, 3).value = pcts[label]
            updated[label] = {"old_pct": old, "new_pct": pcts[label], "row": r}

    return updated
