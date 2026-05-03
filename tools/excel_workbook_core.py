"""Low-level openpyxl helpers shared by all excel_*_core modules.

Pure functions, no @tool decoration. Conventions:
- All functions accept either an open `Workbook` or a `Path`/`str`.
- Save is the caller's responsibility unless the helper is one-shot.
- Named-range writes use workbook scope (None sheet) per BnK template.

Sheet names in BnK_WBS_Template_v1.0.xlsx:
  '0. How to use', '1. Effort', '2. WBS',
  '3. Delivery Plan (By Month)', '3. Delivery Plan',
  '4. Master Data'
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

PathLike = Union[str, Path]

SHEET_EFFORT = "1. Effort"
SHEET_WBS = "2. WBS"
SHEET_DELIVERY = "3. Delivery Plan"
SHEET_DELIVERY_MONTH = "3. Delivery Plan (By Month)"
SHEET_MASTER = "4. Master Data"

# WBS layout (mirrors render_wbs.py)
# NOTE: AI/ML column (I) was added between FE and BA via insert_ai_column_wbs().
WBS_HEADER_ROW = 5
WBS_DATA_START = 9
WBS_COL_INDICATOR = 1   # A
WBS_COL_NUM = 2         # B
WBS_COL_REFCODE = 3     # C
WBS_COL_FEATURE = 4     # D
WBS_COL_DESC = 5        # E
WBS_COL_TOTAL = 6       # F
WBS_COL_BE = 7          # G
WBS_COL_FE = 8          # H
WBS_COL_AI = 9          # I  ← AI/ML effort (new)
WBS_COL_BA = 10         # J  (was I)
WBS_COL_QC = 11         # K  (was J)
WBS_COL_PM = 12         # L  (was K)
WBS_COL_REMARK = 13     # M  (was L)

# Effort layout
EFFORT_HEADER_ROW = 5
EFFORT_DATA_START = 6


def open_wb(path: PathLike) -> Workbook:
    """Load workbook preserving formulas (data_only=False)."""
    return load_workbook(str(path), data_only=False)


def save_wb(wb: Workbook, path: PathLike) -> None:
    wb.save(str(path))


def find_last_data_row(
    ws: Worksheet,
    start_row: int = 1,
    check_cols: tuple[int, ...] = (2, 3, 4),
) -> int:
    """Scan from `ws.max_row` upward, return last row that has a non-None value
    in ANY of `check_cols`. Returns `start_row - 1` if no data found."""
    for r in range(ws.max_row, start_row - 1, -1):
        for c in check_cols:
            v = ws.cell(r, c).value
            if v is not None and (not isinstance(v, str) or v.strip()):
                return r
    return start_row - 1


def clear_row(ws: Worksheet, row: int, max_col: int) -> None:
    """Set every cell in row 1..max_col to None (skipping merged-slave cells)."""
    from openpyxl.cell.cell import MergedCell
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        if not isinstance(cell, MergedCell):
            cell.value = None


def upsert_named_range(wb: Workbook, name: str, sheet: str, cell_ref: str) -> None:
    """Create or replace a workbook-scoped named range pointing to
    `'sheet'!cell_ref` (cell_ref must include $ markers for absolute refs)."""
    value = f"'{sheet}'!{cell_ref}"
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=value)


def col_letter(col: int) -> str:
    return get_column_letter(col)


def cell_has_value(ws: Worksheet, row: int, col: int) -> bool:
    v = ws.cell(row, col).value
    return v is not None and (not isinstance(v, str) or v.strip() != "")
